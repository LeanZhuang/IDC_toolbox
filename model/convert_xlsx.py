import os

import xlrd
from openpyxl import Workbook

from model.IDC_path import idc_path


def convert_xlsx():

    #设置路径
    this_path = idc_path()

    input_path = '/Users/zhuangyuhao/Downloads'

    # 获取最新的xls文件
    xls_files = [os.path.join(input_path, file) for file in os.listdir(input_path) if file.startswith('export')]
    latest_file = max(xls_files, key=os.path.getctime)

    # 打开XLS文件
    xls_workbook = xlrd.open_workbook(latest_file)

    # 创建一个新的XLSX文件
    xlsx_workbook = Workbook()

    # 获取XLS文件的第一个工作表
    xls_sheet = xls_workbook.sheet_by_index(0)

    # 创建一个对应的XLSX工作表
    xlsx_sheet = xlsx_workbook.active

    # 逐行逐列复制数据
    for row_index in range(xls_sheet.nrows):
        for col_index in range(xls_sheet.ncols):
            cell_value = xls_sheet.cell(row_index, col_index).value
            xlsx_sheet.cell(row=row_index+1, column=col_index+1).value = cell_value

    # 设置保存路径
    save_directory = this_path + "/temp"
    save_file_path = save_directory + "/中间底稿.xlsx"

    # 保存XLSX文件
    xlsx_workbook.save(save_file_path)
