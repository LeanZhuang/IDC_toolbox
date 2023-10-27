from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from model.formatted import formatted_accured_data, formatted_bandwith, formatted_no_bandwith
import os
from model.IDC_path import idc_path


def generate_base():
    this_path = idc_path()
    output_path = '/Users/zhuangyuhao/Desktop'

    # 加载底稿
    base_wb = formatted_accured_data()
    base_ws = base_wb.active

    # 加载匹配到的预提
    wb1 = formatted_no_bandwith()
    ws1 = wb1.active
    ws1.insert_rows(0)
    ws1.insert_rows(0)

    wb2 = formatted_bandwith()
    ws2 = wb2.active
    ws2.insert_rows(0)
    ws2.insert_rows(0)

    # 合并三个Excel文件的数据
    merged_data = list(base_ws.values) + list(ws1.values) + list(ws2.values)

    # 创建空工作簿
    merged_wb = load_workbook(this_path + '/temp/empty_file.xlsx')

    # 设置工作表为合并后数据的工作表
    merged_ws = merged_wb.active
    merged_ws.title = "Merged Data"


    # 添加合并后的数据
    for row in merged_data:
        merged_ws.append(row)


    last_row = merged_ws.max_row
    row = last_row + 1
    cell = f'E{row}'
    merged_ws[cell] = f'=(D{row}-C{row})/D{row}'

    cell = f'F{row}'
    merged_ws[cell] = f'=AVERAGE(C{row},D{row})'

    cell = f'I{row}'
    merged_ws[cell] = f'=G{row}*H{row}'


    row = base_ws.max_row + ws1.max_row
    cell = f'R{row}'
    merged_ws[cell] = f'=SUM(R{base_ws.max_row + 4}:R{row - 1})'


    # 配置颜色
    blue = PatternFill(fill_type='solid', fgColor='ACD6FF')
    gray = PatternFill(fill_type='solid', fgColor='DFDFDF')


    for row in range(1, 2):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M','N','O','P','Q','R','S','T','U','V','W']:
            cell = '{}{}'.format(col, row)
            merged_ws[cell].fill = blue

    for row in range(base_ws.max_row - 3, base_ws.max_row + 1):
        for col in ['E', 'F', 'G']:
            cell = '{}{}'.format(col, row)
            merged_ws[cell].fill = gray

    for row in range(base_ws.max_row + 3, base_ws.max_row + 4):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y']:
            cell = '{}{}'.format(col, row)
            merged_ws[cell].fill = blue

    for row in range(base_ws.max_row + 3 + ws1.max_row, base_ws.max_row + 4 + ws1.max_row):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R','S','T','U','V','W','X','Y','Z','AA','AB','AC']:
            cell = '{}{}'.format(col, row)
            merged_ws[cell].fill = blue

    for row in range(base_ws.max_row + ws1.max_row + ws2.max_row, base_ws.max_row + ws1.max_row + ws2.max_row + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            cell = '{}{}'.format(col, row)
            merged_ws[cell].fill = gray


    # 保存合并后的Excel文件

    merged_wb.save(output_path + '/【底稿】.xlsx')

# generate_base()