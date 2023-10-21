from openpyxl import load_workbook
from model.IDC_path import idc_path
import os

def formatted_accured_data():
    this_path = idc_path()

    # 加载工作簿
    workbook = load_workbook(this_path + '/temp/中间底稿.xlsx')
    sheet = workbook.active

    # 定义值和公式所在的区域范围
    last_row = sheet.max_row

    word_list = ['非带宽小计', '带宽小计', '合计', '付款申请与对账单核对']

    formula_list1 = [f'=SUMIF(M1:M{last_row},"机架",F1:F{last_row})+SUMIF(M1:M{last_row},"IP",F1:F{last_row})',
                    f'=SUMIF(M1:M{last_row},"端口组",F1:F{last_row})',
                    f'=F{last_row + 1}+F{last_row + 2}']

    formula_list2 = [f'=F{last_row + 1}-VLOOKUP("合计",Q:R,2,FALSE)',
                    f'=F{last_row + 2}-VLOOKUP("合计",H:I,2,FALSE)']


    # 循环遍历要输入值的列表
    for index, value in enumerate(word_list):
        row = last_row + index + 1  # 计算要输入行的行号
        cell = 'E{}'.format(row)  # 构建单元格地址
        sheet[cell] = value  # 输入值到单元格

    # 循环遍历要输入公式的范围每一行
    for index, value in enumerate(formula_list1):
        row = last_row + index + 1  # 计算要输入行的行号
        cell = 'F{}'.format(row)  # 构建单元格地址
        sheet[cell] = value  # 输入值到单元格

    # 循环遍历要输入公式的范围每一行
    for index, value in enumerate(formula_list2):
        row = last_row + index + 1  # 计算要输入行的行号
        cell = 'G{}'.format(row)  # 构建单元格地址
        sheet[cell] = value  # 输入值到单元格


    # 保存工作簿
    return workbook


# 格式化带宽预提
def formatted_bandwith():
    # 加载工作簿
    this_path = idc_path()
    workbook = load_workbook(this_path + '/temp/带宽.xlsx')
    sheet = workbook.active

    # 定义值和公式所在的区域范围
    last_row = sheet.max_row

    col_list = {'A':'地点', 'B':'地点', 'C':'SYS统计',
                'D':'运营商统计', 'E':'差异率', 'F':'中值',
                'G':'结算流量', 'H':'计费单位', 'I':'结算'}


    # 循环遍历要输入值的列表
    for key, value in col_list.items():
        row = last_row + 2  # 计算要输入行的行号
        cell = f'{key}{row}'  # 构建单元格地址
        sheet[cell] = value  # 输入值到单元格

    return workbook



# 格式化非带宽预提
def formatted_no_bandwith():
    # 加载工作簿
    this_path = idc_path()
    workbook = load_workbook(this_path + '/temp/非带宽.xlsx')
    sheet = workbook.active

    # 定义值和公式所在的区域范围
    last_row = sheet.max_row

    row = last_row + 1
    cell = f'R{row}'
    sheet[cell] = f'=SUM(R2:R{row-1})'

    cell = f'Q{row}'
    sheet[cell] = f'合计'

    return workbook

formatted_accured_data()